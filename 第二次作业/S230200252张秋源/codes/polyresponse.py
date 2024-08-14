import numpy as np
import io
import numbers
def getpoly1(varleft : int, nmleft : int, lastDivide : list[int], arr : list[list[int]]) -> None:
    '''
    Get the coefficient of certain polynominal at certain power
    This function is for recursive finding porpose
    ------------------
    For example
    getpoly1(3, 4, [], arr)
    print(arr)
    [[0, 0, 4], [0, 1, 3], [0, 2, 2], [0, 3, 1], [0, 4, 0], 
    [1, 0, 3], [1, 1, 2], [1, 2, 1], [1, 3, 0], [2, 0, 2], [2, 1, 1], 
    [2, 2, 0], [3, 0, 1], [3, 1, 0], [4, 0, 0]]
    '''
    if(varleft <= 1):
        lastDivide.append(nmleft)
        arr.append(lastDivide)
    else:
        v1 = lastDivide.copy()
        for i in range(nmleft + 1):
            v1.append(i)
            getpoly1(varleft - 1, nmleft - i, v1.copy(), arr)
            v1 = lastDivide.copy()

#arr = []
#getpoly1(3, 4, [], arr)
#print(arr)

def getallpoly(varinum : int, expon : int) -> list[list[int]]:
    '''
    Get exponent all possible polynominals, given the number of variables and
    the maximum exponent
    '''
    out = []
    for i in range(expon + 1):
        tmp = []
        getpoly1(varinum, i, [], tmp)
        out += tmp
    return out

#print(getallpoly(3, 4))

def getpartpoly(varinum : int, expon : int, crossExpon : int) -> list[list[int]]:
    '''
    Get exponent of all possible polynominals up to crossExpon, the cross exponents above
    will not be considered
    '''
    if(crossExpon > expon):
        raise ValueError("You can't make cross exponent large than required exponent")
    out = []
    for i in range(crossExpon + 1):
        tmp = []
        getpoly1(varinum, i, [], tmp)
        out += tmp
    z1 = [0] * varinum
    for i in range(crossExpon + 1, expon + 1, 1):
        for j in range(varinum):
            t = z1.copy()
            t[j] = i
            out.append(t)
    return out

#print(getpartpoly(3, 4, 4))

def valcal(x : (numbers.Real | np.ndarray), ex : list[int]) -> float:
    '''
    calculate value by given number and exponent
    '''
    if(isinstance(x, numbers.Real)):
        out = x ** ex[0]
        return out
    elif(type(x) == np.ndarray):
        lx = x.shape[0]
        if(lx != len(ex)):
            raise ValueError("x and ex size mismatch")
        out = 1.0
        for i in range(lx):
            out *= (x[i] ** ex[i])
        return out

#print(valcal([7, 8], [2, 3]))

class PolyResponse:
    def __init__(self, xvals : (list[list[float]] | np.ndarray), yvals : (list[float] | np.ndarray), expon : int, crossExpon : int):
        #if(len(xvals) != len(yvals)):
            #raise RuntimeError("Invalid variable input")
        self.xvalsarr : np.ndarray
        self.yvalsarr : np.ndarray
        if(type(xvals) == list):
            xpointnum = len(xvals)
            self.varinum = len(xvals[0]) #number of dimensions
        elif(type(xvals) == np.ndarray):
            xpointnum = xvals.shape[0]
            if(len(xvals.shape) == 1):
                self.varinum = 1
            else:
                self.varinum = xvals.shape[1]
        else:
            raise ValueError("xvals invalid")
        if(type(yvals) == list):
            ypointnum = len(yvals)
        elif(type(xvals) == np.ndarray):
            ypointnum = yvals.shape[0]
        else:
            raise ValueError("yvals invalid")
        if(xpointnum != ypointnum):
            raise RuntimeError("Invalid variable input")
        self.xvalsarr = np.array(xvals)
        self.yvalsarr = np.array(yvals)
        self.poly = getpartpoly(self.varinum, expon, crossExpon)
        self.coeff = calcoeff(self.xvalsarr, self.yvalsarr, self.poly)
        #print(self.coeff)
    def __call__(self, xval : (list[float] | np.ndarray | numbers.Real)) -> float:
        out = 0
        for i in range(len(self.poly)):
            out += valcal(xval, self.poly[i]) * self.coeff[i]
        return out
    def express(self) -> str:
        out = io.StringIO()
        for i in range(len(self.coeff)):
            out.write(f'({self.coeff[i]})' )
            if(len(self.poly[i]) != 1): # multiple X variable
                for j in range(len(self.poly[i])):
                    if(self.poly[i][j] != 0):
                        out.write(f" * x[{j}] ** {self.poly[i][j]}")
                if(i != len(self.coeff) - 1):
                    out.write(" + ")
            else: # One X variable
                for j in range(len(self.poly[i])):
                    if(self.poly[i][j] != 0):
                        out.write(f" * x ** {self.poly[i][j]}")
                if(i != len(self.coeff) - 1):
                    out.write(" + ")
        return out.getvalue()
    
    def to_excel(self, excelfname : str) -> None:
        import openpyxl
        wb = openpyxl.Workbook()
        wb.create_sheet("Data")
        ws = wb["Data"]
        ws.cell(1, 1).value = "Coefficient"
        ws.cell(1, 2).value = "Value"
        ws.cell(1, 3).value = "Coefficient"
        ws.cell(1, 4).value = "Value"
        for i in range(len(self.coeff)):
            if(len(self.poly[i]) != 1):
                s1 = ""
                for j in range(len(self.poly[i])):
                    if(self.poly[i][j] != 0):
                        s1 += "x_{" + f"{j + 1}" + "}^{" + f"{self.poly[i][j]}" + "}"
                if(s1 == ""):
                    s1 = "1"
            else:
                if(self.poly[i][0] != 0):
                    s1 = "x^{" + f"{self.poly[i][0]}" + "}"
                else:
                    s1 = "1"
            rownum = i // 2 + 2
            colnum = (i % 2) * 2 + 1
            ws.cell(rownum, colnum).value = s1
            ws.cell(rownum, colnum + 1).value = self.coeff[i]
        wb.save(excelfname)

class PolyResponse2:
    '''
    First transfer them to mean = 0, variance = 1, then use polynominal
    response surface
    '''
    def __init__(self, xvals : (list[list[float]] | np.ndarray), yvals : (list[float] | np.ndarray), expon : int, crossExpon : int):
        #if(len(xvals) != len(yvals)):
            #raise RuntimeError("Invalid variable input")
        self.xvalsarr : np.ndarray
        self.yvalsarr : np.ndarray
        if(type(xvals) == list):
            xpointnum = len(xvals)
            self.varinum = len(xvals[0]) #number of dimensions
        elif(type(xvals) == np.ndarray):
            xpointnum = xvals.shape[0]
            if(len(xvals.shape) == 1):
                self.varinum = 1
            else:
                self.varinum = xvals.shape[1]
        else:
            raise ValueError("xvals invalid")
        if(type(yvals) == list):
            ypointnum = len(yvals)
        elif(type(xvals) == np.ndarray):
            ypointnum = yvals.shape[0]
        else:
            raise ValueError("yvals invalid")
        if(xpointnum != ypointnum):
            raise RuntimeError("Invalid variable input")
        self.xvalsarr = np.array(xvals)
        self.yvalsarr = np.array(yvals)

        self.xmean = np.mean(self.xvalsarr, axis=0)
        self.xstd = np.std(self.xvalsarr, axis=0)
        if(len(self.xvalsarr.shape) == 1): # x 1 dimensional
            self.xvalsarr -= self.xmean
            self.xvalsarr /= self.xstd
        else: # x multiple dimensional
            for i in range(self.xvalsarr.shape[1]):
                self.xvalsarr[:, i] -= self.xmean[i]
                self.xvalsarr[:, i] /= self.xstd[i]
        self.poly = getpartpoly(self.varinum, expon, crossExpon)
        self.coeff = calcoeff(self.xvalsarr, self.yvalsarr, self.poly)
        #print(self.coeff)
    def __call__(self, xval : (list[float] | np.ndarray | numbers.Real)) -> float:
        out = 0
        if(isinstance(xval, numbers.Real)):
            testx = xval
            testx -= self.xmean
            testx /= self.xstd
        else:
            testx = np.array(xval)
            for i in range(testx.shape[0]):
                testx[i] -= self.xmean[i]
                testx[i] /= self.xstd[i]
        for i in range(len(self.poly)):
            out += valcal(testx, self.poly[i]) * self.coeff[i]
        return out
    def express(self) -> str:
        out = io.StringIO()
        # Get previous process expression
        if(len(self.poly[0]) == 1): # One dimension
            out.write(f"x -= {self.xmean}\n")
            out.write(f"x /= {self.xstd}\n")
        else: # Multiple dimension
            self.xmean : np.ndarray
            for i in range(self.xmean.shape[0]):
                out.write(f"x[{i}] -= {self.xmean[i]}\n")
                out.write(f"x[{i}] /= {self.xstd[i]}\n")
        for i in range(len(self.coeff)):
            out.write(f'({self.coeff[i]})' )
            if(len(self.poly[i]) != 1):
                for j in range(len(self.poly[i])):
                    if(self.poly[i][j] != 0):
                        out.write(f" * x[{j}] ** {self.poly[i][j]}")
                if(i != len(self.coeff) - 1):
                    out.write(" + ")
            else:
                for j in range(len(self.poly[i])):
                    if(self.poly[i][j] != 0):
                        out.write(f" * x ** {self.poly[i][j]}")
                if(i != len(self.coeff) - 1):
                    out.write(" + ")
        return out.getvalue()
    
    def express2(self) -> str:
        out = io.StringIO()
        # Get previous process expression
        if(len(self.poly[0]) == 1): # One dimension
            out.write(f"x -= {self.xmean}\n")
            out.write(f"x /= {self.xstd}\n")
        else: # Multiple dimension
            self.xmean : np.ndarray
            for i in range(self.xmean.shape[0]):
                out.write(f"x_{i + 1} -= {self.xmean[i]}\n")
                out.write(f"x_{i + 1} /= {self.xstd[i]}\n")
        for i in range(len(self.coeff)):
            out.write(f'({self.coeff[i]:.4e})' )
            if(len(self.poly[i]) != 1):
                for j in range(len(self.poly[i])):
                    if(self.poly[i][j] > 1):
                        out.write(f"*((x_{j + 1})^{self.poly[i][j]})")
                    elif(self.poly[i][j] == 1):
                        out.write(f"*(x_{j + 1})")
                    else:
                        pass
                if(i != len(self.coeff) - 1):
                    out.write("+")
            else:
                for j in range(len(self.poly[i])):
                    if(self.poly[i][j] != 0):
                        out.write(f"*x^{self.poly[i][j]}")
                if(i != len(self.coeff) - 1):
                    out.write("+")
        return out.getvalue()
    
    def to_latex(self) -> str:
        out = io.StringIO()
        # Get previous process expression
        if(len(self.poly[0]) == 1): # One dimension
            out.write(f"x -= {self.xmean}\n")
            out.write(f"x /= {self.xstd}\n")
        else: # Multiple dimension
            self.xmean : np.ndarray
            for i in range(self.xmean.shape[0]):
                out.write("x_{" + str(i + 1) + "} -= " + f"{self.xmean[i]}\n")
                out.write("x_{" + str(i + 1) + "} /= " + f"{self.xstd[i]}\n")
        for i in range(len(self.coeff)):
            #out.write(f'({self.coeff[i]})' )
            skiptag : bool = False
            if(i != 0):
                if(self.coeff[i] > 0):
                    out.write("+")
                elif(self.coeff[i] < 0):
                    out.write("-")
                else:
                    skiptag = True
            if(not skiptag):
                if(i != 0):
                    out.write(f"{abs(self.coeff[i]):1.3f}")
                else:
                    out.write(f"{self.coeff[i]:1.3f}")
                if(len(self.poly[i]) != 1):
                    for j in range(len(self.poly[i])):
                        if(self.poly[i][j] != 0):
                            #out.write(f" * x[{j}] ** {self.poly[i][j]}")
                            out.write("x_{" + str(j + 1) + "}")
                            if(self.poly[i][j] != 1):
                                out.write("^{" + str(self.poly[i][j]) + "} ")
                    #if(i != len(self.coeff) - 1):
                        #out.write(" + ")
                else:
                    for j in range(len(self.poly[i])):
                        if(self.poly[i][j] != 0):
                            #out.write(f" * x ** {self.poly[i][j]}")
                            out.write("x")
                            if(self.poly[i][j] != 1):
                                out.write("^{" + str(self.poly[i][j]) + "} ")
                    #if(i != len(self.coeff) - 1):
                        #out.write(" + ")
        return out.getvalue()
    
    def to_excel(self, excelfname : str) -> None:
        import openpyxl
        wb = openpyxl.Workbook()
        wb.create_sheet("Data")
        ws = wb["Data"]
        ws.cell(1, 1).value = "Coefficient"
        ws.cell(1, 2).value = "Value"
        ws.cell(1, 3).value = "Coefficient"
        ws.cell(1, 4).value = "Value"
        for i in range(len(self.coeff)):
            if(len(self.poly[i]) != 1):
                s1 = ""
                for j in range(len(self.poly[i])):
                    if(self.poly[i][j] != 0):
                        s1 += "x_{" + f"{j + 1}" + "}^{" + f"{self.poly[i][j]}" + "}"
                if(s1 == ""):
                    s1 = "1"
            else:
                if(self.poly[i][0] != 0):
                    s1 = "x^{" + f"{self.poly[i][0]}" + "}"
                else:
                    s1 = "1"
            rownum = i // 2 + 2
            colnum = (i % 2) * 2 + 1
            ws.cell(rownum, colnum).value = s1
            ws.cell(rownum, colnum + 1).value = self.coeff[i]
        wb.save(excelfname)



def calcoeff(xvals : np.ndarray, yvals : np.ndarray, poly : list[list[int]]) -> np.ndarray:
    '''
    Calculate coefficient of polynominal
    '''
    n = xvals.shape[0] # The number of points
    m = len(poly) # The number of coefficients
    phimat = np.zeros((n, m), dtype=np.float64)
    for i in range(n):
        for j in range(m):
            phimat[i, j] = valcal(xvals[i], poly[j])
    tpphimat = phimat.transpose() #reference
    leftmat = np.matmul(tpphimat, phimat)
    for i in range(n):
        tpphimat[:, i] *= yvals[i]
    rightmat = np.sum(tpphimat, axis=1)
    out = np.linalg.solve(leftmat, rightmat)
    #out = np.matmul(np.linalg.inv(leftmat), rightmat)
    #print(xvals)
    #print(phimat)
    #print(xvals[0])
    return out

'''
xvals = [[2.2, 0.17, 0.012], [3.3, 0.30, 0.033], [2.6, 0.18, 0.041], [2.9, 0.10, 0.023], [2.3, 0.26, 0.059]]
yvals = [0.296, 0.289, 0.285, 0.282, 0.273]
a = PolyResponse(xvals, yvals, 2, 2)
print(a.express())
'''

'''
xvals = np.array([[1,2], [3,5], [7,10], [13,6], [8,7], [4,5]])
yvals = np.array([2.3, 4.5, 6.8, 7.4, 8.2, 11.2])
a = PolyResponse(xvals, yvals, 2, 2)
print(a.express())
for i in xvals:
    print(a.f(i))
'''


'''
xvals = np.array([1, 3, 7, 13, 8, 4])
yvals = np.array([2.3, 4.5, 6.8, 7.4, 8.2, 11.2])
a = PolyResponse(xvals, yvals, 2, 2)
print(a.express())
for i in xvals:
    print(a.f(i))
'''



